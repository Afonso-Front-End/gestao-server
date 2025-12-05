"""
Rotas para geração de relatórios Excel - D1
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from typing import Optional
from datetime import datetime
import logging
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.database import get_database
from app.core.collections import COLLECTION_D1_BIPAGENS

logger = logging.getLogger(__name__)

router = APIRouter(tags=["D-1 - Relatórios"])


@router.get("/bipagens/gerar-relatorio-contato")
async def gerar_relatorio_contato_d1(
    base: Optional[str] = Query(None, description="Lista de bases separadas por vírgula"),
    tempo_parado: Optional[str] = Query(None, description="Lista de tempos parados separados por vírgula"),
    cidade: Optional[str] = Query(None, description="Lista de cidades separadas por vírgula")
):
    """
    📊 GERA RELATÓRIO EXCEL DE CONTATO COM MOTORISTAS D1
    Retorna arquivo Excel com: Base, Nome do Motorista, Total, Total Entregue, Total Não Entregue, Status, Observação
    """
    try:
        db = get_database()
        if db is None:
            raise HTTPException(status_code=500, detail="Database não está conectado")
        
        # Normalizar filtros
        bases_list = [b.strip() for b in base.split(',')] if base else []
        tempos_list = [t.strip() for t in tempo_parado.split(',')] if tempo_parado else []
        cidades_list = [c.strip() for c in cidade.split(',')] if cidade else []
        
        collection = db[COLLECTION_D1_BIPAGENS]
        
        # Construir query de match
        match_query = {}
        if bases_list:
            match_query['$or'] = [
                {'base_entrega': {'$in': bases_list}},
                {'base_escaneamento': {'$in': bases_list}}
            ]
        
        if tempos_list:
            match_query['tempo_pedido_parado'] = {'$in': tempos_list}
        
        if cidades_list:
            match_query['cidade_destino'] = {'$in': cidades_list}
        
        # Pipeline para buscar motoristas agrupados (mesma lógica do endpoint /motoristas)
        pipeline = [
            {'$match': match_query} if match_query else {'$match': {}},
            {'$sort': {
                'numero_pedido_jms': 1,
                'tempo_digitalizacao': -1
            }},
            {'$group': {
                '_id': '$numero_pedido_jms',
                'doc': {'$first': '$$ROOT'}
            }},
            {'$replaceRoot': {'newRoot': '$doc'}},
            {'$match': {
                'responsavel_entrega': {'$exists': True, '$ne': '', '$ne': None},
                'esta_com_motorista': True
            }},
            {'$group': {
                '_id': '$responsavel_entrega',
                'base_entrega': {'$first': '$base_entrega'},
                'total_pedidos': {'$sum': 1},
                'pedidos': {'$push': {
                    'marca_assinatura': '$marca_assinatura'
                }}
            }},
            {'$addFields': {
                'total_entregues': {
                    '$size': {
                        '$filter': {
                            'input': '$pedidos',
                            'as': 'pedido',
                            'cond': {
                                '$or': [
                                    {'$regexMatch': {'input': {'$toLower': {'$ifNull': ['$$pedido.marca_assinatura', '']}}, 'regex': 'assinatura de devolução|recebimento com assinatura normal'}},
                                    {'$eq': [{'$toLower': {'$ifNull': ['$$pedido.marca_assinatura', '']}}, 'entregue']}
                                ]
                            }
                        }
                    }
                },
                'total_nao_entregues': {
                    '$size': {
                        '$filter': {
                            'input': '$pedidos',
                            'as': 'pedido',
                            'cond': {
                                '$regexMatch': {'input': {'$toLower': {'$ifNull': ['$$pedido.marca_assinatura', '']}}, 'regex': 'não entregue|nao entregue'}
                            }
                        }
                    }
                }
            }},
            {'$project': {
                'motorista': '$_id',
                'base_entrega': 1,
                'total_pedidos': 1,
                'total_entregues': 1,
                'total_nao_entregues': 1,
                '_id': 0
            }},
            {'$sort': {'base_entrega': 1, 'motorista': 1}}
        ]
        
        stats = {}
        async for doc in collection.aggregate(pipeline):
            motorista = doc.get('motorista', '')
            base_entrega = doc.get('base_entrega', '')
            key_motorista = f"{motorista}||{base_entrega}" if base_entrega else motorista
            
            stats[key_motorista] = {
                "motorista": motorista,
                "base": base_entrega,
                "total": doc.get('total_pedidos', 0),
                "entregues": doc.get('total_entregues', 0),
                "nao_entregues": doc.get('total_nao_entregues', 0)
            }
        
        # Buscar status e observações dos motoristas D1
        motoristas_status_collection = db["motoristas_status_d1"]
        status_map = {}
        observacoes_map = {}
        
        for key_motorista, data in stats.items():
            motorista = data["motorista"]
            base = data["base"]
            
            if base:
                status_doc = await motoristas_status_collection.find_one({
                    "responsavel": motorista,
                    "base": base
                })
            else:
                status_doc = await motoristas_status_collection.find_one({
                    "responsavel": motorista,
                    "$or": [
                        {"base": {"$exists": False}},
                        {"base": None},
                        {"base": ""}
                    ]
                })
            
            status_map[key_motorista] = status_doc.get("status", "") if status_doc else ""
            observacoes_map[key_motorista] = status_doc.get("observacao", "") if status_doc else ""
        
        # Criar arquivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Contato D1"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalhos
        headers = ["Base", "Nome do Motorista", "Total", "Total Entregue", "Total Não Entregue", "Status", "Observação"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Dados
        data_list = list(stats.values())
        data_list.sort(key=lambda x: (x["base"] or "", x["motorista"]))
        
        for row_idx, data in enumerate(data_list, start=2):
            key_motorista = f"{data['motorista']}||{data['base']}" if data['base'] else data['motorista']
            status = status_map.get(key_motorista, "")
            observacao = observacoes_map.get(key_motorista, "")
            
            ws.cell(row=row_idx, column=1, value=data["base"] or "N/A").border = border
            ws.cell(row=row_idx, column=2, value=data["motorista"]).border = border
            ws.cell(row=row_idx, column=3, value=data["total"]).border = border
            ws.cell(row=row_idx, column=4, value=data["entregues"]).border = border
            ws.cell(row=row_idx, column=5, value=data["nao_entregues"]).border = border
            ws.cell(row=row_idx, column=6, value=status).border = border
            ws.cell(row=row_idx, column=7, value=observacao).border = border
            
            # Alinhar números ao centro
            for col in [3, 4, 5]:
                ws.cell(row=row_idx, column=col).alignment = center_alignment
            
            # Alinhar observação à esquerda (texto longo)
            ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 50  # Coluna de Observação
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Converter para bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gerar nome do arquivo
        agora = datetime.now()
        data_formatada = agora.strftime("%Y%m%d")
        hora_formatada = agora.strftime("%H%M%S")
        
        if bases_list:
            base_nome = re.sub(r'[<>:"/\\|?*]', '_', bases_list[0]).strip()
            base_nome = re.sub(r'\s+', '_', base_nome)
            
            if len(bases_list) > 1:
                base_nome_completo = f"{base_nome}_e_{len(bases_list)-1}_outras"
            else:
                base_nome_completo = base_nome
            
            filename = f"Relatorio_Contato_D1_{base_nome_completo}_{data_formatada}_{hora_formatada}.xlsx"
        else:
            filename = f"Relatorio_Contato_D1_Todas_Bases_{data_formatada}_{hora_formatada}.xlsx"
        
        logger.info(f"✅ Relatório D1 gerado: {filename} com {len(data_list)} motoristas")
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório D1: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")

