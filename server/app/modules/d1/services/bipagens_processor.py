"""
Processador para dados de bipagens em tempo real
"""
import openpyxl
from io import BytesIO
from typing import List, Dict, Any
import logging
import re
from datetime import datetime, timedelta
from app.services.database import get_database
from app.core.collections import COLLECTION_D1_CHUNKS, COLLECTION_D1_BIPAGENS
from bson import ObjectId

logger = logging.getLogger(__name__)


class BipagensProcessor:
    """
    Processador para dados de bipagens em tempo real
    - Deduplica por número de pedido (pega data mais recente)
    - Valida e cruza dados com d1_chunks
    - Calcula tempo de pedido parado
    """
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls']
    
    async def process_file(
        self, 
        file_content: bytes, 
        filename: str
    ) -> Dict[str, Any]:
        """
        Processa arquivo Excel de bipagens em tempo real
        
        Args:
            file_content: Conteúdo do arquivo em bytes
            filename: Nome do arquivo
            
        Returns:
            Dict com resultado do processamento
        """
        try:
            # Verificar formato
            if not any(filename.lower().endswith(fmt) for fmt in self.supported_formats):
                raise ValueError(f"Formato não suportado. Use: {', '.join(self.supported_formats)}")
            
            # Ler Excel
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            sheet = workbook.active
            
            # Ler cabeçalhos
            headers = [cell.value for cell in sheet[1]]
            
            # Validar colunas necessárias
            required_columns = [
                'Número de pedido JMS',
                'Tempo de digitalização',
                'Correio de coleta ou entrega',
                'Tipo de bipagem'
            ]
            
            # Colunas opcionais mas importantes
            optional_columns = ['Digitalizador', 'Base Destino', 'Base de escaneamento']
            
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise ValueError(f"Colunas obrigatórias não encontradas: {', '.join(missing_columns)}")
            
            # Verificar se colunas opcionais existem
            has_digitalizador = 'Digitalizador' in headers
            has_base_destino = 'Base Destino' in headers
            has_base_escaneamento = 'Base de escaneamento' in headers
            
            if not has_digitalizador:
                logger.warning("⚠️ Coluna 'Digitalizador' não encontrada. Pedidos sem motorista não serão processados corretamente.")
            if not has_base_destino and not has_base_escaneamento:
                logger.warning("⚠️ Colunas 'Base Destino' ou 'Base de escaneamento' não encontradas. Validação de base não será aplicada.")
            
            # Ler dados de forma otimizada (processar em lotes para grandes arquivos)
            dados_brutos = []
            batch_size = 10000  # Processar em lotes de 10k linhas
            total_rows = sheet.max_row - 1  # Excluir cabeçalho
            
            logger.info(f"📊 Iniciando leitura de {total_rows} linhas do arquivo...")
            
            for start_row in range(2, sheet.max_row + 1, batch_size):
                end_row = min(start_row + batch_size, sheet.max_row + 1)
                batch_data = []
                
                for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
                    if not row[0]:  # Pular linhas vazias
                        continue
                    
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        if header:
                            row_dict[str(header)] = row[idx] if idx < len(row) else None
                    
                    if row_dict.get('Número de pedido JMS'):
                        batch_data.append(row_dict)
                
                dados_brutos.extend(batch_data)
                logger.info(f"📊 Lidas {len(dados_brutos)}/{total_rows} linhas...")
            
            logger.info(f"📊 Total de linhas lidas: {len(dados_brutos)}")
            
            # Deduplicar por número de pedido (pegar data mais recente)
            dados_deduplicados = self._deduplicar_por_data_recente(dados_brutos)
            logger.info(f"📊 Total após deduplicação: {len(dados_deduplicados)}")
            
            # Buscar dados completos na coleção d1_chunks
            dados_completos = await self._buscar_dados_completos(dados_deduplicados)
            logger.info(f"📊 Total de pedidos com dados completos: {len(dados_completos)}")
            
            # Calcular tempo de pedido parado
            dados_finais = self._calcular_tempo_parado(dados_completos)
            
            # Salvar na nova coleção
            resultado = await self._salvar_na_colecao(dados_finais)
            
            return {
                "success": True,
                "total_linhas_lidas": len(dados_brutos),
                "total_deduplicados": len(dados_deduplicados),
                "total_processados": len(dados_finais),
                "total_salvos": resultado["saved"],
                "total_atualizados": resultado["updated"],
                "filename": filename
            }
            
        except Exception as e:
            logger.error(f"Erro ao processar arquivo de bipagens: {e}", exc_info=True)
            raise
    
    def _deduplicar_por_data_recente(self, dados: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Deduplica pedidos mantendo o melhor registro:
        - Prioriza registros com "Correio de coleta ou entrega" preenchido
        - Se o último bipe não tiver motorista, busca o último bipe que tenha motorista
        - Se todos tiverem motorista, mantém o mais recente
        """
        # Agrupar todos os registros por número de pedido
        pedidos_agrupados = {}
        
        for item in dados:
            numero_pedido = str(item.get('Número de pedido JMS', '')).strip()
            if not numero_pedido:
                continue
            
            # Remover pedidos filhos (formato: 888001152307637-001, 888001152307637-002, etc.)
            is_child = bool(
                re.search(r"\.\d+$", numero_pedido) or 
                re.search(r"-\d+$", numero_pedido) or 
                re.search(r"_\d+$", numero_pedido) or 
                re.search(r"[A-Za-z]$", numero_pedido)
            )
            if is_child:
                continue  # Pular pedidos filhos
            
            tempo_digitalizacao_str = item.get('Tempo de digitalização')
            if not tempo_digitalizacao_str:
                continue
            
            # Converter para datetime
            try:
                if isinstance(tempo_digitalizacao_str, datetime):
                    tempo_digitalizacao = tempo_digitalizacao_str
                elif isinstance(tempo_digitalizacao_str, str):
                    # Tentar vários formatos
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                        try:
                            tempo_digitalizacao = datetime.strptime(tempo_digitalizacao_str, fmt)
                            break
                        except:
                            continue
                    else:
                        continue
                else:
                    continue
            except:
                continue
            
            # Adicionar ao grupo de pedidos
            if numero_pedido not in pedidos_agrupados:
                pedidos_agrupados[numero_pedido] = []
            
            item['_tempo_digitalizacao'] = tempo_digitalizacao
            pedidos_agrupados[numero_pedido].append(item)
        
        # Para cada pedido, escolher o registro mais recente (independente de ter motorista ou não)
        # REGRA: Sempre usar a bipagem mais recente. Se ela não tem motorista, o pedido não está com motorista.
        pedidos_finais = []
        
        for numero_pedido, registros in pedidos_agrupados.items():
            # Ordenar por data (mais recente primeiro)
            registros.sort(key=lambda x: x['_tempo_digitalizacao'], reverse=True)
            
            # SEMPRE pegar o registro mais recente (independente de ter motorista ou não)
            melhor_registro = registros[0]
            tipo_bipagem = melhor_registro.get('Tipo de bipagem', '')
            correio = str(melhor_registro.get('Correio de coleta ou entrega', '')).strip()
            tem_motorista = correio and correio != ''
            
            if tem_motorista:
                logger.info(f"✅ Pedido {numero_pedido}: Bipagem mais recente COM MOTORISTA - Tipo: {tipo_bipagem}, Motorista: {correio}, Data: {melhor_registro['_tempo_digitalizacao']}")
                melhor_registro['_tem_motorista'] = True
            else:
                logger.info(f"📦 Pedido {numero_pedido}: Bipagem mais recente SEM MOTORISTA - Tipo: {tipo_bipagem}, Data: {melhor_registro['_tempo_digitalizacao']}")
                logger.info(f"   ℹ️ Pedido não está com motorista (entrou no galpão ou não foi atribuído)")
                melhor_registro['_tem_motorista'] = False
            
            # Remover campo auxiliar
            melhor_registro.pop('_tempo_digitalizacao', None)
            pedidos_finais.append(melhor_registro)
        
        return pedidos_finais
    
    async def _buscar_dados_completos(self, dados_bipagens: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Busca dados completos na coleção d1_chunks e mescla com dados de bipagens
        Otimizado para grandes volumes usando aggregation pipeline
        """
        db = get_database()
        collection_chunks = db[COLLECTION_D1_CHUNKS]
        
        # Extrair números de pedidos únicos
        numeros_pedidos = [str(item.get('Número de pedido JMS', '')).strip() for item in dados_bipagens if item.get('Número de pedido JMS')]
        numeros_pedidos = list(set(numeros_pedidos))
        
        logger.info(f"🔍 Buscando {len(numeros_pedidos)} pedidos únicos em d1_chunks...")
        
        # Criar um dicionário para armazenar pedidos encontrados
        pedidos_encontrados = {}
        numeros_pedidos_set = set(numeros_pedidos)  # Usar set para busca O(1)
        
        # Contador para log
        chunks_processados = 0
        total_chunks = await collection_chunks.count_documents({})
        
        logger.info(f"🔍 Varrendo {total_chunks} chunks em busca de {len(numeros_pedidos)} pedidos...")
        
        # Varrer todos os chunks uma única vez (mais eficiente)
        async for chunk_doc in collection_chunks.find({}):
            chunks_processados += 1
            if chunks_processados % 100 == 0:
                logger.info(f"🔍 Processados {chunks_processados}/{total_chunks} chunks... ({len(pedidos_encontrados)} pedidos encontrados)")
            
            chunk_data = chunk_doc.get('chunk_data', [])
            if isinstance(chunk_data, list):
                for registro in chunk_data:
                    numero_pedido = str(registro.get('Número de pedido JMS', '')).strip()
                    # Verificar se está na lista de pedidos procurados e ainda não foi encontrado
                    if numero_pedido in numeros_pedidos_set and numero_pedido not in pedidos_encontrados:
                        pedidos_encontrados[numero_pedido] = registro
                        
                        # Se já encontrou todos, pode parar
                        if len(pedidos_encontrados) >= len(numeros_pedidos):
                            logger.info(f"✅ Todos os {len(numeros_pedidos)} pedidos encontrados!")
                            break
                
                # Se já encontrou todos, parar de varrer chunks
                if len(pedidos_encontrados) >= len(numeros_pedidos):
                    break
        
        logger.info(f"✅ Encontrados {len(pedidos_encontrados)}/{len(numeros_pedidos)} pedidos em d1_chunks ({len(pedidos_encontrados)*100//len(numeros_pedidos) if numeros_pedidos else 0}%)")
        
        dados_completos = []
        
        for item_bipagem in dados_bipagens:
            numero_pedido = str(item_bipagem.get('Número de pedido JMS', '')).strip()
            if not numero_pedido:
                continue
            
            # Remover pedidos filhos (formato: 888001152307637-001, 888001152307637-002, etc.)
            # Verificar se é pedido filho (tem hífen seguido de números, ponto seguido de números, ou letra no final)
            is_child = bool(
                re.search(r"\.\d+$", numero_pedido) or 
                re.search(r"-\d+$", numero_pedido) or 
                re.search(r"_\d+$", numero_pedido) or 
                re.search(r"[A-Za-z]$", numero_pedido)
            )
            if is_child:
                continue  # Pular pedidos filhos
            
            # Verificar se "Correio de coleta ou entrega" está preenchido
            correio = str(item_bipagem.get('Correio de coleta ou entrega', '')).strip()
            correio_vazio = not correio or correio == '' or correio.strip() == ''
            
            # Buscar dados do d1_chunks para validações extras
            pedido_encontrado = pedidos_encontrados.get(numero_pedido)
            
            if not pedido_encontrado:
                logger.warning(f"⚠️ Pedido {numero_pedido} não encontrado em d1_chunks")
                continue
            
            # Dados do d1_chunks
            responsavel_entrega_chunks = str(pedido_encontrado.get('Responsável pela entrega', '')).strip()
            base_entrega_chunks = str(pedido_encontrado.get('Base de entrega', '')).strip()
            
            # Dados do arquivo de bipagens
            digitalizador = str(item_bipagem.get('Digitalizador', '')).strip()
            tipo_bipagem = str(item_bipagem.get('Tipo de bipagem', '')).strip()
            base_destino = str(item_bipagem.get('Base Destino', '')).strip()
            base_escaneamento = str(item_bipagem.get('Base de escaneamento', '')).strip()
            
            # Usar "Base de escaneamento" se disponível, senão usar "Base Destino"
            base_escaneamento_final = base_escaneamento if base_escaneamento else base_destino
            
            # VALIDAÇÃO 1: Verificar se bases são iguais
            bases_iguais = False
            if base_escaneamento_final and base_entrega_chunks:
                base_escaneamento_normalizada = base_escaneamento_final.strip().upper()
                base_entrega_normalizada = base_entrega_chunks.strip().upper()
                bases_iguais = base_escaneamento_normalizada == base_entrega_normalizada
            
            # VALIDAÇÃO 2: Verificar se "Digitalizador" = "Responsável pela entrega" (do d1_chunks)
            digitalizador_igual_responsavel = False
            if digitalizador and responsavel_entrega_chunks:
                digitalizador_normalizado = digitalizador.strip().upper()
                responsavel_normalizado = responsavel_entrega_chunks.strip().upper()
                digitalizador_igual_responsavel = digitalizador_normalizado == responsavel_normalizado
            
            # VALIDAÇÃO 3: Verificar se "Tipo de bipagem" = "bipe de pacote problemático"
            tipo_bipagem_problematico = False
            if tipo_bipagem:
                tipo_bipagem_normalizado = tipo_bipagem.strip().upper()
                tipo_bipagem_problematico = 'bipe de pacote problemático' in tipo_bipagem_normalizado or 'pacote problemático' in tipo_bipagem_normalizado
            
            # Determinar se pedido está com motorista:
            # 1. "Correio de coleta ou entrega" preenchido E bases iguais
            # 2. OU "Digitalizador" = "Responsável pela entrega" E "Correio" vazio E bases iguais
            # 3. OU "Tipo de bipagem" = "bipe de pacote problemático" E bases iguais
            esta_com_motorista = False
            responsavel_final = ''
            
            if bases_iguais:
                if not correio_vazio:
                    # Caso 1: Correio preenchido
                    esta_com_motorista = True
                    responsavel_final = correio
                    logger.info(f"📋 Pedido {numero_pedido} COM MOTORISTA (Correio preenchido) - Motorista: {correio}, Base: {base_entrega_chunks}")
                elif digitalizador_igual_responsavel:
                    # Caso 2: Digitalizador = Responsável pela entrega
                    esta_com_motorista = True
                    responsavel_final = responsavel_entrega_chunks
                    logger.info(f"📋 Pedido {numero_pedido} COM MOTORISTA (Digitalizador = Responsável) - Motorista: {responsavel_entrega_chunks}, Base: {base_entrega_chunks}")
                elif tipo_bipagem_problematico:
                    # Caso 3: Tipo de bipagem = "bipe de pacote problemático"
                    esta_com_motorista = True
                    # Usar responsável do d1_chunks se disponível, senão usar digitalizador
                    responsavel_final = responsavel_entrega_chunks if responsavel_entrega_chunks else digitalizador
                    logger.info(f"📋 Pedido {numero_pedido} COM MOTORISTA (Bipe de pacote problemático) - Motorista: {responsavel_final}, Base: {base_entrega_chunks}")
                else:
                    # Sem motorista
                    logger.info(f"📦 Pedido {numero_pedido} SEM MOTORISTA - Base escaneamento: {base_escaneamento_final}, Base entrega: {base_entrega_chunks}")
            else:
                # Bases diferentes = sem motorista
                logger.info(f"📦 Pedido {numero_pedido} SEM MOTORISTA (bases diferentes) - Base escaneamento: {base_escaneamento_final}, Base entrega: {base_entrega_chunks}")
            
            tempo_digitalizacao = item_bipagem.get('Tempo de digitalização')
            
            # Definir base_final (usar base_entrega_chunks, com fallback para base_escaneamento_final)
            base_final = base_entrega_chunks if base_entrega_chunks else base_escaneamento_final
            
            # Mesclar dados
            # Atualizar "Correio de coleta ou entrega" com o responsavel_final (pode ser do correio, digitalizador ou d1_chunks)
            correio_final = responsavel_final if esta_com_motorista else correio
            
            dados_finais = {
                'Número de pedido JMS': numero_pedido,
                'Base de entrega': base_final,
                'Horário de saída para entrega': pedido_encontrado.get('Horário de saída para entrega', ''),
                'Responsável pela entrega': responsavel_final,
                'Marca de assinatura': pedido_encontrado.get('Marca de assinatura', ''),
                'CEP destino': pedido_encontrado.get('CEP destino', ''),
                'Motivos dos pacotes problemáticos': pedido_encontrado.get('Motivos dos pacotes problemáticos', ''),
                'Destinatário': pedido_encontrado.get('Destinatário', ''),
                'Complemento': pedido_encontrado.get('Complemento', ''),
                'Distrito destinatário': pedido_encontrado.get('Distrito destinatário', ''),
                'Cidade Destino': pedido_encontrado.get('Cidade Destino', ''),
                '3 Segmentos': pedido_encontrado.get('3 Segmentos', ''),
                'Tempo de digitalização': tempo_digitalizacao,
                'Correio de coleta ou entrega': correio_final,  # Usar responsavel_final se tiver motorista
                'Tipo de bipagem': tipo_bipagem,
                'Digitalizador': digitalizador,
                'Base Destino': base_escaneamento_final,
                'Base de escaneamento': base_escaneamento_final,
                '_esta_com_motorista': esta_com_motorista,  # Flag auxiliar para uso no salvamento
                '_responsavel_entrega_chunks_original': responsavel_entrega_chunks  # Salvar original do d1_chunks para validações no fallback
            }
            
            dados_completos.append(dados_finais)
        
        return dados_completos
    
    def _calcular_tempo_parado(self, dados: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Calcula o tempo de pedido parado baseado na data de digitalização
        """
        hoje = datetime.now()
        
        for item in dados:
            tempo_digitalizacao_str = item.get('Tempo de digitalização')
            if not tempo_digitalizacao_str:
                item['Tempo de Pedido parado'] = None
                continue
            
            # Converter para datetime
            try:
                if isinstance(tempo_digitalizacao_str, datetime):
                    tempo_digitalizacao = tempo_digitalizacao_str
                elif isinstance(tempo_digitalizacao_str, str):
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                        try:
                            tempo_digitalizacao = datetime.strptime(tempo_digitalizacao_str, fmt)
                            break
                        except:
                            continue
                    else:
                        item['Tempo de Pedido parado'] = None
                        continue
                else:
                    item['Tempo de Pedido parado'] = None
                    continue
                
                # Calcular diferença em dias
                diferenca = hoje - tempo_digitalizacao
                dias = diferenca.days
                
                # Formatar como "Exceed X days with no track"
                item['Tempo de Pedido parado'] = f"Exceed {dias} days with no track"
                
            except Exception as e:
                logger.warning(f"Erro ao calcular tempo parado: {e}")
                item['Tempo de Pedido parado'] = None
        
        return dados
    
    async def _salvar_na_colecao(self, dados: List[Dict[str, Any]]) -> Dict[str, int]:
        """
        Salva ou atualiza dados na coleção d1_bipagens usando bulk operations para performance
        """
        from pymongo import UpdateOne
        
        db = get_database()
        collection = db[COLLECTION_D1_BIPAGENS]
        
        # Criar índice único em numero_pedido_jms se não existir
        try:
            await collection.create_index('numero_pedido_jms', unique=True, background=True)
        except Exception as e:
            logger.debug(f"Índice já existe ou erro ao criar: {e}")
        
        saved = 0
        updated = 0
        
        # Preparar operações bulk
        bulk_operations = []
        hoje = datetime.now()
        
        # Processar em chunks de 1000 para não sobrecarregar memória
        chunk_size = 1000
        total_chunks = (len(dados) + chunk_size - 1) // chunk_size
        
        for chunk_idx in range(0, len(dados), chunk_size):
            chunk = dados[chunk_idx:chunk_idx + chunk_size]
            logger.info(f"💾 Processando chunk {chunk_idx // chunk_size + 1}/{total_chunks} ({len(chunk)} registros)")
            
            bulk_operations = []
            
            for item in chunk:
                numero_pedido = str(item.get('Número de pedido JMS', '')).strip()
                if not numero_pedido:
                    continue
                
                # Remover pedidos filhos (formato: 888001152307637-001, 888001152307637-002, etc.)
                # Verificar se é pedido filho (tem hífen seguido de números, ponto seguido de números, ou letra no final)
                is_child = bool(
                    re.search(r"\.\d+$", numero_pedido) or 
                    re.search(r"-\d+$", numero_pedido) or 
                    re.search(r"_\d+$", numero_pedido) or 
                    re.search(r"[A-Za-z]$", numero_pedido)
                )
                if is_child:
                    continue  # Pular pedidos filhos
                
                # Converter tempo de digitalização para datetime se necessário
                tempo_digitalizacao = item.get('Tempo de digitalização')
                if isinstance(tempo_digitalizacao, str):
                    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                        try:
                            tempo_digitalizacao = datetime.strptime(tempo_digitalizacao, fmt)
                            break
                        except:
                            continue
                
                # Usar flag _esta_com_motorista se disponível (já calculada no processamento)
                # Caso contrário, recalcular usando a mesma lógica
                if '_esta_com_motorista' in item:
                    esta_com_motorista_final = item['_esta_com_motorista']
                    responsavel_entrega_valor = item.get('Responsável pela entrega', '')
                else:
                    # Fallback: recalcular validação
                    correio_valor = item.get('Correio de coleta ou entrega', '')
                    correio_preenchido = bool(correio_valor and str(correio_valor).strip() != '')
                    
                    # VALIDAÇÃO ROBUSTA: Verificar se Base de escaneamento é igual a Base de entrega
                    base_entrega_valor = item.get('Base de entrega', '')
                    base_escaneamento_valor = item.get('Base de escaneamento', '') or item.get('Base Destino', '')
                    
                    bases_iguais = True
                    if base_escaneamento_valor and base_entrega_valor:
                        base_escaneamento_normalizada = str(base_escaneamento_valor).strip().upper()
                        base_entrega_normalizada = str(base_entrega_valor).strip().upper()
                        bases_iguais = base_escaneamento_normalizada == base_entrega_normalizada
                    
                    # Se as bases forem diferentes, considerar como SEM MOTORISTA
                    esta_com_motorista_final = correio_preenchido and bases_iguais
                    
                    # Se "Correio de coleta ou entrega" está vazio OU bases são diferentes, responsavel_entrega fica vazio
                    responsavel_entrega_valor = str(correio_valor).strip() if esta_com_motorista_final else ''
                
                documento = {
                    'numero_pedido_jms': numero_pedido,
                    'base_entrega': item.get('Base de entrega', ''),
                    'horario_saida_entrega': item.get('Horário de saída para entrega', ''),
                    'responsavel_entrega': responsavel_entrega_valor,  # Vazio se não tem motorista
                    'marca_assinatura': item.get('Marca de assinatura', ''),
                    'cep_destino': item.get('CEP destino', ''),
                    'motivos_pacotes_problematicos': item.get('Motivos dos pacotes problemáticos', ''),
                    'destinatario': item.get('Destinatário', ''),
                    'complemento': item.get('Complemento', ''),
                    'distrito_destinatario': item.get('Distrito destinatário', ''),
                    'cidade_destino': item.get('Cidade Destino', ''),
                    'tres_segmentos': item.get('3 Segmentos', ''),
                    'tempo_digitalizacao': tempo_digitalizacao,
                    'tempo_pedido_parado': item.get('Tempo de Pedido parado', ''),
                    'digitalizador': item.get('Digitalizador', ''),
                    'base_destino': item.get('Base Destino', ''),
                    'base_escaneamento': item.get('Base de escaneamento', '') or item.get('Base Destino', ''),
                    'esta_com_motorista': esta_com_motorista_final,  # True apenas se "Correio" preenchido E bases iguais
                    'updated_at': hoje
                }
                
                # Usar upsert (insere se não existe, atualiza se existe)
                operation = UpdateOne(
                    {'numero_pedido_jms': numero_pedido},
                    {
                        '$set': documento,
                        '$setOnInsert': {'created_at': hoje}
                    },
                    upsert=True
                )
                bulk_operations.append(operation)
            
            # Executar bulk operations
            if bulk_operations:
                try:
                    result = await collection.bulk_write(bulk_operations, ordered=False)
                    saved += result.upserted_count
                    updated += result.modified_count
                    logger.info(f"✅ Chunk processado: {result.upserted_count} inseridos, {result.modified_count} atualizados")
                except Exception as e:
                    logger.error(f"❌ Erro ao processar chunk: {e}")
                    # Fallback: processar um por um
                    for item in chunk:
                        try:
                            numero_pedido = str(item.get('Número de pedido JMS', '')).strip()
                            if not numero_pedido:
                                continue
                            
                            # Converter tempo de digitalização
                            tempo_digitalizacao = item.get('Tempo de digitalização')
                            if isinstance(tempo_digitalizacao, str):
                                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y']:
                                    try:
                                        tempo_digitalizacao = datetime.strptime(tempo_digitalizacao, fmt)
                                        break
                                    except:
                                        continue
                            
                            # Usar flag _esta_com_motorista se disponível (já calculada no processamento)
                            if '_esta_com_motorista' in item:
                                esta_com_motorista_fallback = item['_esta_com_motorista']
                                responsavel_entrega_fallback = item.get('Responsável pela entrega', '')
                            else:
                                # Fallback: recalcular validação completa
                                correio_valor_fallback = item.get('Correio de coleta ou entrega', '')
                                correio_vazio_fallback = not correio_valor_fallback or str(correio_valor_fallback).strip() == ''
                                
                                base_entrega_valor_fallback = item.get('Base de entrega', '')
                                base_escaneamento_valor_fallback = item.get('Base de escaneamento', '') or item.get('Base Destino', '')
                                digitalizador_fallback = str(item.get('Digitalizador', '')).strip()
                                tipo_bipagem_fallback = str(item.get('Tipo de bipagem', '')).strip()
                                # Usar o responsável original do d1_chunks se disponível, senão usar o processado
                                responsavel_entrega_chunks_fallback = str(item.get('_responsavel_entrega_chunks_original', '') or item.get('Responsável pela entrega', '')).strip()
                                
                                # VALIDAÇÃO 1: Verificar se bases são iguais
                                bases_iguais_fallback = False
                                if base_escaneamento_valor_fallback and base_entrega_valor_fallback:
                                    base_escaneamento_normalizada_fallback = str(base_escaneamento_valor_fallback).strip().upper()
                                    base_entrega_normalizada_fallback = str(base_entrega_valor_fallback).strip().upper()
                                    bases_iguais_fallback = base_escaneamento_normalizada_fallback == base_entrega_normalizada_fallback
                                
                                # VALIDAÇÃO 2: Verificar se "Digitalizador" = "Responsável pela entrega"
                                digitalizador_igual_responsavel_fallback = False
                                if digitalizador_fallback and responsavel_entrega_chunks_fallback:
                                    digitalizador_normalizado_fallback = digitalizador_fallback.strip().upper()
                                    responsavel_normalizado_fallback = responsavel_entrega_chunks_fallback.strip().upper()
                                    digitalizador_igual_responsavel_fallback = digitalizador_normalizado_fallback == responsavel_normalizado_fallback
                                
                                # VALIDAÇÃO 3: Verificar se "Tipo de bipagem" = "bipe de pacote problemático"
                                tipo_bipagem_problematico_fallback = False
                                if tipo_bipagem_fallback:
                                    tipo_bipagem_normalizado_fallback = tipo_bipagem_fallback.strip().upper()
                                    tipo_bipagem_problematico_fallback = 'bipe de pacote problemático' in tipo_bipagem_normalizado_fallback or 'pacote problemático' in tipo_bipagem_normalizado_fallback
                                
                                # Determinar se está com motorista (mesma lógica do processamento principal)
                                esta_com_motorista_fallback = False
                                if bases_iguais_fallback:
                                    if not correio_vazio_fallback:
                                        esta_com_motorista_fallback = True
                                        responsavel_entrega_fallback = str(correio_valor_fallback).strip()
                                    elif digitalizador_igual_responsavel_fallback:
                                        esta_com_motorista_fallback = True
                                        responsavel_entrega_fallback = responsavel_entrega_chunks_fallback
                                    elif tipo_bipagem_problematico_fallback:
                                        esta_com_motorista_fallback = True
                                        responsavel_entrega_fallback = responsavel_entrega_chunks_fallback if responsavel_entrega_chunks_fallback else digitalizador_fallback
                                else:
                                    responsavel_entrega_fallback = ''
                            
                            documento = {
                                'numero_pedido_jms': numero_pedido,
                                'base_entrega': item.get('Base de entrega', ''),
                                'horario_saida_entrega': item.get('Horário de saída para entrega', ''),
                                'responsavel_entrega': responsavel_entrega_fallback,
                                'marca_assinatura': item.get('Marca de assinatura', ''),
                                'cep_destino': item.get('CEP destino', ''),
                                'motivos_pacotes_problematicos': item.get('Motivos dos pacotes problemáticos', ''),
                                'destinatario': item.get('Destinatário', ''),
                                'complemento': item.get('Complemento', ''),
                                'distrito_destinatario': item.get('Distrito destinatário', ''),
                                'cidade_destino': item.get('Cidade Destino', ''),
                                'tres_segmentos': item.get('3 Segmentos', ''),
                                'tempo_digitalizacao': tempo_digitalizacao,
                                'tempo_pedido_parado': item.get('Tempo de Pedido parado', ''),
                                'digitalizador': item.get('Digitalizador', ''),
                                'base_destino': item.get('Base Destino', ''),
                                'base_escaneamento': item.get('Base de escaneamento', '') or item.get('Base Destino', ''),
                                'esta_com_motorista': esta_com_motorista_fallback,
                                'updated_at': hoje
                            }
                            
                            existing = await collection.find_one({'numero_pedido_jms': numero_pedido})
                            if existing:
                                await collection.update_one({'numero_pedido_jms': numero_pedido}, {'$set': documento})
                                updated += 1
                            else:
                                documento['created_at'] = hoje
                                await collection.insert_one(documento)
                                saved += 1
                        except Exception as e2:
                            logger.warning(f"⚠️ Erro ao processar item individual: {e2}")
                            continue
        
        return {
            "saved": saved,
            "updated": updated
        }

