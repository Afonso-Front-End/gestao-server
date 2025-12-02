"""
Processador otimizado para D-1 - grandes volumes (200k-400k registros)
"""
import openpyxl
from io import BytesIO
from typing import List, Dict, Any, Tuple
import logging
from concurrent.futures import ThreadPoolExecutor
import asyncio

logger = logging.getLogger(__name__)


class D1Processor:
    """
    Processador otimizado para grandes volumes de dados D-1
    - Processamento assíncrono
    - Chunks otimizados
    - Memória eficiente
    """
    
    def __init__(self, chunk_size: int = 5000):
        """
        Inicializa o processador
        
        Args:
            chunk_size: Tamanho do chunk (padrão 5000 para grandes volumes)
        """
        self.chunk_size = chunk_size
        self.supported_formats = ['.xlsx', '.xls']
    
    async def process_file(
        self, 
        file_content: bytes, 
        filename: str
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Processa arquivo Excel de forma otimizada para grandes volumes
        
        Args:
            file_content: Conteúdo do arquivo em bytes
            filename: Nome do arquivo
            
        Returns:
            Tuple com (dados_processados, colunas_encontradas)
        """
        try:
            # Verificar formato
            if not any(filename.lower().endswith(fmt) for fmt in self.supported_formats):
                raise ValueError(f"Formato não suportado. Use: {', '.join(self.supported_formats)}")
            
            logger.info(f"📊 Iniciando processamento de {filename} (otimizado para grandes volumes)")
            
            # Processar em thread separada para não bloquear
            loop = asyncio.get_event_loop()
            with ThreadPoolExecutor() as executor:
                data, columns = await loop.run_in_executor(
                    executor,
                    self._process_excel_sync,
                    file_content,
                    filename
                )
            
            logger.info(f"✅ Processamento concluído: {len(data)} registros, {len(columns)} colunas")
            
            return data, columns
            
        except Exception as e:
            logger.error(f"❌ Erro ao processar arquivo: {str(e)}")
            raise
    
    def _process_excel_sync(self, file_content: bytes, filename: str) -> Tuple[List[Dict[str, Any]], List[str]]:
        """
        Processa Excel de forma síncrona (executado em thread separada)
        """
        # Tentar ler arquivo Excel - começar com modo normal para garantir compatibilidade
        # read_only pode ter problemas com alguns arquivos
        try:
            # Primeiro tentar modo normal (mais compatível)
            workbook = openpyxl.load_workbook(
                BytesIO(file_content),
                data_only=True
            )
            logger.info("📖 Arquivo aberto em modo normal")
        except Exception as e:
            logger.warning(f"⚠️ Erro ao abrir em modo normal: {e}. Tentando read_only...")
            # Fallback: modo read_only se normal falhar
            try:
                workbook = openpyxl.load_workbook(
                    BytesIO(file_content),
                    read_only=True,
                    data_only=True
                )
                logger.info("📖 Arquivo aberto em modo read_only")
            except Exception as e2:
                logger.error(f"❌ Erro ao abrir arquivo em ambos os modos: {e2}")
                raise
        
        sheet = workbook.active
        logger.info(f"📄 Planilha ativa: {sheet.title}, {sheet.max_row} linhas, {sheet.max_column} colunas")
        
        # Obter cabeçalhos da primeira linha
        headers = []
        try:
            # Tentar ler primeira linha com values_only=True (mais rápido)
            first_row_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row_values:
                for header_value in first_row_values:
                    if header_value is None or str(header_value).strip() == '':
                        headers.append(f"col_{len(headers)}")
                    else:
                        headers.append(str(header_value).strip())
            else:
                # Fallback: ler com cells
                first_row = list(sheet.iter_rows(min_row=1, max_row=1))[0] if sheet.max_row > 0 else []
                for cell in first_row:
                    header_value = cell.value if hasattr(cell, 'value') else None
                    if header_value is None or str(header_value).strip() == '':
                        headers.append(f"col_{len(headers)}")
                    else:
                        headers.append(str(header_value).strip())
        except Exception as e:
            logger.error(f"Erro ao ler cabeçalhos: {e}")
            # Criar cabeçalhos genéricos baseados no número de colunas
            num_cols = sheet.max_column if hasattr(sheet, 'max_column') else 0
            headers = [f"col_{i}" for i in range(num_cols)]
        
        logger.info(f"📋 Cabeçalhos encontrados: {len(headers)} colunas")
        if len(headers) > 0:
            logger.info(f"   Primeiras 5 colunas: {headers[:5]}")
        
        # Processar linhas de forma otimizada
        data = []
        row_count = 0
        empty_row_count = 0
        
        logger.info(f"📋 Processando linhas do arquivo (cabeçalhos: {len(headers)} colunas, max_row: {sheet.max_row})...")
        
        # Verificar se há linhas para processar
        if sheet.max_row <= 1:
            logger.warning(f"⚠️ Arquivo parece ter apenas cabeçalhos (max_row: {sheet.max_row})")
            workbook.close()
            return data, headers
        
        # Processar primeiras linhas para debug
        sample_rows_processed = 0
        
        # Usar iter_rows para processar linha por linha (economiza memória)
        try:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                non_null_count = 0
                
                # Processar cada célula da linha
                for i, value in enumerate(row):
                    if i < len(headers):
                        processed_value = self._process_cell_value(value)
                        row_dict[headers[i]] = processed_value
                        # Contar valores não-nulos
                        if processed_value is not None:
                            non_null_count += 1
                
                # Adicionar linha se tiver pelo menos UM dado válido (mais permissivo)
                # Isso garante que mesmo linhas com poucos dados sejam incluídas
                if non_null_count > 0:
                    data.append(row_dict)
                    # Log das primeiras 5 linhas processadas
                    if len(data) <= 5:
                        logger.info(f"📝 Linha {row_idx} processada: {non_null_count}/{len(headers)} valores não-nulos")
                        # Mostrar amostra dos valores
                        sample_values = {k: str(v)[:30] if v else 'None' for k, v in list(row_dict.items())[:3]}
                        logger.info(f"   Amostra: {sample_values}")
                else:
                    empty_row_count += 1
                    # Log das primeiras 3 linhas vazias para debug
                    if empty_row_count <= 3:
                        logger.debug(f"⚠️ Linha {row_idx} vazia: {len(row)} células, valores: {[str(v)[:20] if v else 'None' for v in list(row)[:5]]}")
                
                row_count += 1
                
                # Log de progresso a cada 50k registros
                if row_count % 50000 == 0:
                    logger.info(f"📈 Processados {row_count:,} linhas ({len(data):,} com dados, {empty_row_count:,} vazias)...")
                
                # Log detalhado das primeiras 10 linhas
                if row_count <= 10:
                    logger.info(f"🔍 Linha {row_idx}: {non_null_count} valores não-nulos de {len(headers)} colunas")
        except Exception as e:
            logger.error(f"❌ Erro ao processar linhas: {e}")
            logger.exception(e)
        
        workbook.close()
        
        logger.info(f"✅ Processamento concluído: {row_count:,} linhas totais, {len(data):,} linhas com dados, {empty_row_count:,} linhas vazias")
        
        if len(data) == 0 and row_count > 0:
            logger.warning(f"⚠️ ATENÇÃO: {row_count} linhas foram lidas mas nenhuma tinha dados válidos!")
            logger.warning(f"   Verifique se o arquivo tem dados além dos cabeçalhos")
        
        return data, headers
    
    def _process_cell_value(self, value: Any) -> Any:
        """
        Processa valor de célula de forma otimizada
        """
        if value is None:
            return None
        
        # Se for datetime, converter para string ISO
        try:
            if hasattr(value, 'isoformat'):
                return value.isoformat()
            if hasattr(value, 'strftime'):
                return value.strftime('%Y-%m-%d %H:%M:%S')
        except:
            pass
        
        # Se for número, manter como número (mas converter int grande para string)
        if isinstance(value, (int, float)):
            # Números muito grandes (como códigos de pedido) manter como string
            if isinstance(value, int) and abs(value) > 999999999:
                return str(value)
            return value
        
        # Converter para string e limpar
        try:
            str_value = str(value).strip()
            # Retornar vazio como None para economizar espaço
            if str_value == '':
                return None
            return str_value
        except Exception as e:
            logger.warning(f"Erro ao processar valor da célula: {e}, valor: {value}")
            return str(value) if value is not None else None
    
    def create_chunks(self, data: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
        """
        Divide dados em chunks otimizados
        
        Args:
            data: Lista de dados
            
        Returns:
            Lista de chunks
        """
        chunks = []
        for i in range(0, len(data), self.chunk_size):
            chunk = data[i:i + self.chunk_size]
            chunks.append(chunk)
        
        logger.info(f"📦 Dados divididos em {len(chunks)} chunks de até {self.chunk_size} registros")
        
        return chunks

