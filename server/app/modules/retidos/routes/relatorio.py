"""
Rotas para geração de relatórios Excel
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from typing import List, Optional
from datetime import datetime
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.database import get_database
from app.core.collections import COLLECTION_PEDIDOS_RETIDOS_TABELA_CHUNKS
from .helpers import (
    get_numero_pedido,
    get_base_entrega,
    get_responsavel,
    get_marca_assinatura,
    is_child_pedido,
    is_entregue,
    is_nao_entregue,
    extract_raiz_numero
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Pedidos Retidos - Relatórios"])


@router.get("/gerar-relatorio-contato")
async def gerar_relatorio_contato(
    bases: Optional[str] = Query(None, description="Lista de bases separadas por vírgula")
):
    """
    📊 GERA RELATÓRIO EXCEL DE CONTATO COM MOTORISTAS
    Retorna arquivo Excel com: Base, Nome do Motorista, Total, Total Entregue, Total Não Entregue, Status, Observação
    """
    try:
        db = get_database()
        if db is None:
            raise HTTPException(status_code=500, detail="Database não está conectado")
        
        # Normalizar filtros de bases
        bases_list = [b.strip() for b in bases.split(',')] if bases else []
        
        # Buscar dados dos pedidos parados
        import re
        
        collection = db[COLLECTION_PEDIDOS_RETIDOS_TABELA_CHUNKS]
        stats = {}
        raiz_vistas = set()
        total_validos = 0
        
        cursor = collection.find({}).sort("chunk_number", 1)
        async for chunk in cursor:
            for item in chunk.get("data", []) or []:
                # Filtrar por bases se fornecido
                if bases_list:
                    base_item = get_base_entrega(item)
                    if not base_item or base_item not in bases_list:
                        continue
                
                # Extrair e validar número do pedido
                numero = get_numero_pedido(item)
                if not numero:
                    continue
                
                # Remover pedidos filhos
                if is_child_pedido(numero):
                    continue
                
                # Deduplicar por raiz numérica
                raiz = extract_raiz_numero(numero)
                if raiz:
                    if raiz in raiz_vistas:
                        continue
                    raiz_vistas.add(raiz)
                
                # Extrair responsável, base e marca
                responsavel = get_responsavel(item)
                base_entrega = get_base_entrega(item)
                marca = get_marca_assinatura(item).lower()
                
                # Criar chave única para motorista+base
                key_motorista = f"{responsavel}||{base_entrega}" if base_entrega else responsavel
                
                if key_motorista not in stats:
                    stats[key_motorista] = {
                        "responsavel": responsavel,
                        "base": base_entrega,
                        "total": 0,
                        "entregues": 0,
                        "nao_entregues": 0,
                    }
                
                # Incrementar contadores
                stats[key_motorista]["total"] += 1
                
                # Classificar status do pedido
                if is_entregue(marca):
                    stats[key_motorista]["entregues"] += 1
                elif is_nao_entregue(marca):
                    stats[key_motorista]["nao_entregues"] += 1
                else:
                    stats[key_motorista]["nao_entregues"] += 1
                
                total_validos += 1
        
        # Buscar status e observações dos motoristas (Pedidos Retidos)
        motoristas_status_collection = db["motoristas_status_pedidos_retidos"]
        status_map = {}
        observacoes_map = {}
        
        for key_motorista, data in stats.items():
            responsavel = data["responsavel"]
            base = data["base"]
            
            if base:
                status_doc = await motoristas_status_collection.find_one({
                    "responsavel": responsavel,
                    "base": base
                })
            else:
                status_doc = await motoristas_status_collection.find_one({
                    "responsavel": responsavel,
                    "$or": [
                        {"base": {"$exists": False}},
                        {"base": None},
                        {"base": ""}
                    ]
                })
            
            status_map[key_motorista] = status_doc.get("status", "") if status_doc else ""
            observacoes_map[key_motorista] = status_doc.get("observacao", "") if status_doc else ""
        
        # Criar arquivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Contato"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalhos
        headers = ["Base", "Nome do Motorista", "Total", "Total Entregue", "Total Não Entregue", "Status", "Observação"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Dados
        data_list = list(stats.values())
        data_list.sort(key=lambda x: (x["base"], x["responsavel"]))
        
        for row_idx, data in enumerate(data_list, start=2):
            key_motorista = f"{data['responsavel']}||{data['base']}" if data['base'] else data['responsavel']
            status = status_map.get(key_motorista, "")
            observacao = observacoes_map.get(key_motorista, "")
            
            ws.cell(row=row_idx, column=1, value=data["base"] or "N/A").border = border
            ws.cell(row=row_idx, column=2, value=data["responsavel"]).border = border
            ws.cell(row=row_idx, column=3, value=data["total"]).border = border
            ws.cell(row=row_idx, column=4, value=data["entregues"]).border = border
            ws.cell(row=row_idx, column=5, value=data["nao_entregues"]).border = border
            ws.cell(row=row_idx, column=6, value=status).border = border
            ws.cell(row=row_idx, column=7, value=observacao).border = border
            
            # Alinhar números ao centro
            for col in [3, 4, 5]:
                ws.cell(row=row_idx, column=col).alignment = center_alignment
            
            # Alinhar observação à esquerda (texto longo)
            ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 50  # Coluna de Observação (mais larga para texto longo)
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Converter para bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gerar nome do arquivo com base, data e hora formatadas
        agora = datetime.now()
        data_formatada = agora.strftime("%Y%m%d")  # Ex: 20250115
        hora_formatada = agora.strftime("%H%M%S")   # Ex: 143022
        
        if bases_list:
            # Limpar e formatar nome da base (remover caracteres inválidos para nome de arquivo)
            # Remover espaços e caracteres especiais, substituir por underscore
            base_nome = re.sub(r'[<>:"/\\|?*]', '_', bases_list[0]).strip()
            base_nome = re.sub(r'\s+', '_', base_nome)  # Substituir espaços por underscore
            
            # Se tiver múltiplas bases, usar o nome da primeira e indicar quantas mais
            if len(bases_list) > 1:
                base_nome_completo = f"{base_nome}_e_{len(bases_list)-1}_outras"
            else:
                base_nome_completo = base_nome
            
            # Formato: Relatorio_Contato_{BASE}_{DATA}_{HORA}.xlsx
            # Exemplo: Relatorio_Contato_BNU_-SC_20250115_143022.xlsx
            filename = f"Relatorio_Contato_{base_nome_completo}_{data_formatada}_{hora_formatada}.xlsx"
        else:
            # Todas as bases: Relatorio_Contato_Todas_Bases_{DATA}_{HORA}.xlsx
            filename = f"Relatorio_Contato_Todas_Bases_{data_formatada}_{hora_formatada}.xlsx"
        
        logger.info(f"✅ Relatório gerado: {filename} com {len(data_list)} motoristas")
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")

