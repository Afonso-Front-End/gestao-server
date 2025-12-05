"""
Rotas para geração de relatórios Excel - SLA
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
from typing import Optional, List
from datetime import datetime
import logging
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.database import get_database
from app.modules.sla.services.sla_calculator import SLACalculator

logger = logging.getLogger(__name__)

router = APIRouter(tags=["SLA - Relatórios"])

# Instância do calculador SLA
sla_calculator = SLACalculator()


@router.get("/gerar-relatorio-contato")
async def gerar_relatorio_contato_sla(
    base: Optional[str] = Query(None, description="Nome da base"),
    cidade: Optional[str] = Query(None, description="Lista de cidades separadas por vírgula")
):
    """
    📊 GERA RELATÓRIO EXCEL DE CONTATO COM MOTORISTAS SLA
    Retorna arquivo Excel com: Base, Nome do Motorista, Total, Total Entregue, Total Não Entregue, % Entregue, Status, Observação
    """
    try:
        if not base:
            raise HTTPException(status_code=400, detail="Parâmetro 'base' é obrigatório")
        
        # Normalizar filtros
        cidades_list = [c.strip() for c in cidade.split(',')] if cidade else None
        
        # Buscar dados dos motoristas usando o SLA calculator
        result = await sla_calculator.calculate_sla_metrics(base, cidades_list)
        
        if not result.get("success") or "motoristas" not in result:
            raise HTTPException(status_code=404, detail="Nenhum dado encontrado para esta base")
        
        motoristas_data = result.get("motoristas", [])
        
        # Buscar status e observações dos motoristas SLA
        db = get_database()
        if db is None:
            raise HTTPException(status_code=500, detail="Database não está conectado")
        
        motoristas_status_collection = db["motorista_status_sla"]
        status_map = {}
        observacoes_map = {}
        
        for motorista_info in motoristas_data:
            motorista = motorista_info.get("motorista", "")
            if not motorista:
                continue
            
            key_motorista = f"{motorista}||{base}"
            
            status_doc = await motoristas_status_collection.find_one({
                "motorista": motorista,
                "base": base
            })
            
            status_map[key_motorista] = status_doc.get("status", "") if status_doc else ""
            observacoes_map[key_motorista] = status_doc.get("observacao", "") if status_doc else ""
        
        # Criar arquivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Contato SLA"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalhos
        headers = ["Base", "Nome do Motorista", "Total", "Total Entregue", "Total Não Entregue", "% Entregue", "Status", "Observação"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Dados
        motoristas_data.sort(key=lambda x: (x.get("motorista", "")))
        
        for row_idx, data in enumerate(motoristas_data, start=2):
            motorista = data.get("motorista", "")
            key_motorista = f"{motorista}||{base}"
            status = status_map.get(key_motorista, "")
            observacao = observacoes_map.get(key_motorista, "")
            
            total = data.get("total", 0)
            entregues = data.get("entregues", 0)
            nao_entregues = data.get("naoEntregues", 0)
            percentual = data.get("percentual_entregues", 0)
            
            ws.cell(row=row_idx, column=1, value=base).border = border
            ws.cell(row=row_idx, column=2, value=motorista).border = border
            ws.cell(row=row_idx, column=3, value=total).border = border
            ws.cell(row=row_idx, column=4, value=entregues).border = border
            ws.cell(row=row_idx, column=5, value=nao_entregues).border = border
            ws.cell(row=row_idx, column=6, value=f"{percentual}%").border = border
            ws.cell(row=row_idx, column=7, value=status).border = border
            ws.cell(row=row_idx, column=8, value=observacao).border = border
            
            # Alinhar números ao centro
            for col in [3, 4, 5, 6]:
                ws.cell(row=row_idx, column=col).alignment = center_alignment
            
            # Alinhar observação à esquerda (texto longo)
            ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 50  # Coluna de Observação
        
        # Congelar primeira linha
        ws.freeze_panes = 'A2'
        
        # Converter para bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gerar nome do arquivo
        agora = datetime.now()
        data_formatada = agora.strftime("%Y%m%d")
        hora_formatada = agora.strftime("%H%M%S")
        
        base_nome = re.sub(r'[<>:"/\\|?*]', '_', base).strip()
        base_nome = re.sub(r'\s+', '_', base_nome)
        
        filename = f"Relatorio_Contato_SLA_{base_nome}_{data_formatada}_{hora_formatada}.xlsx"
        
        logger.info(f"✅ Relatório SLA gerado: {filename} com {len(motoristas_data)} motoristas")
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao gerar relatório SLA: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Erro ao gerar relatório: {str(e)}")

