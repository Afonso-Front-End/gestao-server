"""
Rotas de exportação de dados da lista de telefones
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import Response
import logging
import urllib.parse
import re
from datetime import datetime
from io import BytesIO
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Alignment
from app.core.collections import COLLECTION_TELEFONES
from app.services.database import db

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Lista de Telefones - Exportar"])

def normalizar_hub(hub: str) -> str:
    """
    Normaliza HUB/base para o formato padrão "SIGLA -SC" (com espaço antes do hífen).
    """
    if not hub:
        return hub
    
    hub_original = str(hub).strip().upper()
    
    # Pattern 1: "CD BNU 001" → "BNU -SC"
    match_cd = re.match(r'CD\s+([A-Z]{2,4})\s+\d+', hub_original)
    if match_cd:
        sigla = match_cd.group(1)
        return f"{sigla} -SC"
    
    # Pattern 2: "BNU-SC" ou "ITJ-SC" → "BNU -SC" ou "ITJ -SC"
    match_sem_espaco = re.match(r'^([A-Z]{2,4})-SC$', hub_original)
    if match_sem_espaco:
        sigla = match_sem_espaco.group(1)
        return f"{sigla} -SC"
    
    # Pattern 3: "BNU- SC" ou "ITJ- SC" → "BNU -SC" ou "ITJ -SC"
    match_espaco_depois = re.match(r'^([A-Z]{2,4})-\s+SC$', hub_original)
    if match_espaco_depois:
        sigla = match_espaco_depois.group(1)
        return f"{sigla} -SC"
    
    # Pattern 4: Fix spacing "BNU - SC" → "BNU -SC"
    if re.search(r'\s+-\s+', hub_original):
        return re.sub(r'\s+-\s+', ' -', hub_original)
    
    # Pattern 5: "BNU -SC" (já está correto)
    match_com_espaco = re.match(r'^([A-Z]{2,4})\s+-SC$', hub_original)
    if match_com_espaco:
        sigla = match_com_espaco.group(1)
        return f"{sigla} -SC"
    
    return hub_original

def convert_to_dict(obj):
    """Converte ObjectId para string para serialização JSON"""
    if isinstance(obj, dict):
        result = {}
        for key, value in obj.items():
            if isinstance(value, ObjectId):
                result[key] = str(value)
            elif isinstance(value, dict):
                result[key] = convert_to_dict(value)
            elif isinstance(value, list):
                result[key] = [convert_to_dict(item) if isinstance(item, dict) else item for item in value]
            else:
                result[key] = value
        return result
    return obj

@router.get("/exportar-base/{base}")
async def exportar_base_excel(
    base: str,
    busca: str = Query('', description="Busca opcional por nome do motorista ou telefone")
):
    """
    Exporta os motoristas de uma base específica para Excel
    """
    try:
        base_decoded = urllib.parse.unquote(base)
        base_normalized = normalizar_hub(base_decoded)
        
        collection = db.database[COLLECTION_TELEFONES]
        docs = await collection.find({}).to_list(length=10_000)
        
        itens_achatados = []
        for doc in docs:
            dados_proc = doc.get('dados_processados')
            if isinstance(dados_proc, list) and dados_proc:
                for sub in dados_proc:
                    itens_achatados.append(sub)
            else:
                itens_achatados.append(doc)
        
        motoristas_base = []
        colunas_para_remover = ['_id', 'timestamp', 'origem']
        
        for item in itens_achatados:
            hub = normalizar_hub(str(item.get('HUB', '')))
            if hub == base_normalized:
                item_dict = convert_to_dict(item) if isinstance(item, dict) else item
                for coluna in colunas_para_remover:
                    item_dict.pop(coluna, None)
                motoristas_base.append(item_dict)
        
        if busca:
            busca_lower = busca.lower().strip()
            motoristas_filtrados = []
            for item in motoristas_base:
                try:
                    motorista = str(item.get('Motorista', '')).lower()
                    contato = str(item.get('Contato', '')).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                    busca_sem_formatacao = busca_lower.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
                    
                    if busca_lower in motorista or busca_sem_formatacao in contato:
                        motoristas_filtrados.append(item)
                except Exception as e:
                    logger.warning(f"Erro ao processar item na busca: {e}")
                    continue
            motoristas_base = motoristas_filtrados
        
        motoristas_base.sort(key=lambda x: str(x.get('Motorista', '')).lower())
        
        if not motoristas_base:
            raise HTTPException(status_code=404, detail=f"Nenhum motorista encontrado para a base {base_normalized}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Motoristas"
        
        # Apenas essas colunas serão exportadas
        colunas_necessarias = ['Data', 'Motorista', 'Status', 'Cidade', 'HUB', 'Contato']
        
        # Verificar quais colunas existem nos dados (com variações de nome)
        todas_colunas_disponiveis = set()
        for item in motoristas_base:
            for key in item.keys():
                todas_colunas_disponiveis.add(key)
        
        # Mapear variações de nomes para os nomes padrão
        mapeamento_colunas = {}
        for col_padrao in colunas_necessarias:
            col_padrao_lower = col_padrao.lower()
            for col_disponivel in todas_colunas_disponiveis:
                col_disponivel_lower = col_disponivel.lower()
                # Verificar correspondência exata ou parcial
                if (col_padrao_lower == col_disponivel_lower or 
                    col_padrao_lower in col_disponivel_lower or 
                    col_disponivel_lower in col_padrao_lower):
                    mapeamento_colunas[col_padrao] = col_disponivel
                    break
            
            # Caso especial: "Contato" também pode ser "Telefone"
            if col_padrao == 'Contato' and col_padrao not in mapeamento_colunas:
                for col_disponivel in todas_colunas_disponiveis:
                    col_disponivel_lower = col_disponivel.lower()
                    if 'telefone' in col_disponivel_lower:
                        mapeamento_colunas[col_padrao] = col_disponivel
                        break
        
        # Usar apenas as colunas necessárias (na ordem especificada)
        # Após filtrar, usar os nomes padrão normalizados
        colunas_ordenadas = colunas_necessarias.copy()
        
        # Filtrar os dados para manter apenas as colunas necessárias
        motoristas_base_filtrados = []
        for item in motoristas_base:
            item_filtrado = {}
            for col_padrao, col_mapeada in mapeamento_colunas.items():
                if col_mapeada in item:
                    item_filtrado[col_padrao] = item[col_mapeada]
                elif col_padrao in item:
                    item_filtrado[col_padrao] = item[col_padrao]
            motoristas_base_filtrados.append(item_filtrado)
        
        motoristas_base = motoristas_base_filtrados
        
        # Cabeçalho com formatação básica
        for col_idx, coluna in enumerate(colunas_ordenadas, start=1):
            cell = ws.cell(row=1, column=col_idx, value=coluna)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajustar altura da linha do cabeçalho
        ws.row_dimensions[1].height = 25
        
        # Dados com formatação básica
        for row_idx, item in enumerate(motoristas_base, start=2):
            # Ajustar altura das linhas de dados
            ws.row_dimensions[row_idx].height = 20
            
            for col_idx, coluna in enumerate(colunas_ordenadas, start=1):
                # Usar o nome padrão normalizado (já foi mapeado no filtro)
                valor = item.get(coluna, '')
                if isinstance(valor, dict):
                    valor = str(valor)
                elif valor is None:
                    valor = ''
                else:
                    valor = str(valor)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Ajustar largura das colunas automaticamente
        for col_idx, coluna in enumerate(colunas_ordenadas, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            
            # Calcular largura baseada no conteúdo
            max_length = len(str(coluna))  # Começar com o tamanho do cabeçalho
            
            # Verificar o maior conteúdo da coluna
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            
            # Ajustar largura (adicionar um pouco de espaço extra)
            adjusted_width = min(max_length + 3, 50)  # Máximo de 50 caracteres
            ws.column_dimensions[col_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_nome_arquivo = base_normalized.replace(' ', '_').replace('-', '_')
        filename = f"Lista_Telefones_{base_nome_arquivo}_{data_hora}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao exportar base: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erro interno: {str(e)}")

