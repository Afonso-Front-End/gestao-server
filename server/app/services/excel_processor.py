"""
Excel file processing service
"""
import re
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook
import xlrd
import logging
from app.core.exceptions import FileProcessingException

logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Service for processing Excel files"""
    
    def __init__(self):
        self.supported_extensions = ['.xlsx', '.xls']
    
    async def process_file(
        self,
        file_content: bytes,
        filename: str
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Process Excel file and return data with found columns"""
        try:
            file_extension = self._get_file_extension(filename)
            
            if file_extension == 'xlsx':
                return await self._process_xlsx(file_content)
            elif file_extension == 'xls':
                return await self._process_xls(file_content)
            else:
                raise FileProcessingException(f"Unsupported file format: {file_extension}")
                
        except Exception as e:
            logger.error(f"Error processing file {filename}: {e}")
            raise FileProcessingException(f"Failed to process file: {e}")
    
    def _get_file_extension(self, filename: str) -> str:
        """Get file extension"""
        return filename.lower().split('.')[-1] if '.' in filename else ''
    
    async def _process_xlsx(self, file_content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Process .xlsx file"""
        try:
            from io import BytesIO
            workbook = load_workbook(BytesIO(file_content))
            worksheet = workbook.active
            
            # Get headers
            headers = [cell.value for cell in worksheet[1]]
            headers = [str(header).strip() if header is not None else f"Column_{i+1}" 
                      for i, header in enumerate(headers)]
            
            # Process data rows
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_data = {}
                    for i, cell_value in enumerate(row):
                        if i < len(headers):
                            processed_value = self._process_cell_value(cell_value)
                            row_data[headers[i]] = processed_value
                    data.append(row_data)
            
            logger.info(f"Processed {len(data)} rows from XLSX file")
            return data, headers
            
        except Exception as e:
            logger.error(f"Error processing XLSX file: {e}")
            raise FileProcessingException(f"Failed to process XLSX file: {e}")
    
    async def _process_xls(self, file_content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Process .xls file"""
        try:
            from io import BytesIO
            workbook = xlrd.open_workbook(file_contents=file_content)
            worksheet = workbook.sheet_by_index(0)
            
            # Get headers
            headers = [worksheet.cell_value(0, col) for col in range(worksheet.ncols)]
            headers = [str(header).strip() if header else f"Column_{i+1}" 
                      for i, header in enumerate(headers)]
            
            # Process data rows
            data = []
            for row_idx in range(1, worksheet.nrows):
                row_data = {}
                for col_idx in range(worksheet.ncols):
                    if col_idx < len(headers):
                        cell_value = worksheet.cell_value(row_idx, col_idx)
                        processed_value = self._process_cell_value(cell_value)
                        row_data[headers[col_idx]] = processed_value
                data.append(row_data)
            
            logger.info(f"Processed {len(data)} rows from XLS file")
            return data, headers
            
        except Exception as e:
            logger.error(f"Error processing XLS file: {e}")
            raise FileProcessingException(f"Failed to process XLS file: {e}")
    
    def _process_cell_value(self, value: Any) -> str:
        """Process individual cell value"""
        if value is None:
            return ""
        
        # Convert to string and clean
        str_value = str(value).strip()
        
        # Handle empty strings
        if not str_value or str_value.lower() in ['none', 'null', 'nan']:
            return ""
        
        # Handle numbers (convert large integers to strings to avoid JSON serialization issues)
        if isinstance(value, (int, float)):
            if isinstance(value, int) and value > 2**31 - 1:  # Large integer
                return str(value)
            elif isinstance(value, float) and value.is_integer():
                return str(int(value))
            else:
                return str_value
        
        # Handle dates
        if hasattr(value, 'strftime'):  # datetime object
            return value.strftime('%Y-%m-%d %H:%M:%S')
        
        return str_value
    
    def process_hub_column(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Process HUB column to standardize format"""
                
        for item in data:
            if 'HUB' in item and item['HUB']:
                hub_original = str(item['HUB']).strip()
                hub_processado = hub_original
                
                # Pattern 1: "CD BNU 001" → "BNU -SC"
                match_cd = re.match(r'CD\s+([A-Z]{2,4})\s+\d+', hub_original)
                if match_cd:
                    sigla = match_cd.group(1)
                    hub_processado = f"{sigla} -SC"
                                    
                # Pattern 2: Fix spacing "BNU - SC" → "BNU -SC"
                elif re.search(r'\s+-\s+', hub_original):
                    hub_processado = re.sub(r'\s+-\s+', ' -', hub_original)
                                    
                item['HUB'] = hub_processado
        
        return data